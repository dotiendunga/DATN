import openpyxl
import paho.mqtt.client as paho
from paho import mqtt
import time 
from datetime import datetime
import json
#---------------------------connect to  MQTT----------------------------------

# /*MQTT Broker Connection Details*/
mqtt_broker = "ae501b5ee3194ca682bd67e257459478.s1.eu.hivemq.cloud"
mqtt_username = "RainWay System"
mqtt_password = "012301230123aA#"
mqtt_port = 8883
mqtt_topic = "Train/Location_data"
# data speed receive 

def on_connect(client, userdata, flags, rc, properties=None):
    print("CONNACK received with code %s." % rc)
    # subscribe 1 topic 
    # client.subscribe(mqtt_topic_2)
# Callback function khi nhận được tin nhắn từ MQTT Broker
def on_message(client, userdata, message):
    print("Received message '" 
          + str(message.payload.decode("utf-8")) 
          + "' on topic '"+ message.topic 
          + "' with QoS " + str(message.qos))
# client_id is the given name of the client
client = paho.Client(paho.CallbackAPIVersion.VERSION2)
# connect to MQTT
client.on_connect = on_connect

# enable TLS for secure connection
client.tls_set(tls_version=mqtt.client.ssl.PROTOCOL_TLS)
# set username and password
client.username_pw_set(mqtt_username,mqtt_password)
# connect to HiveMQ Cloud on port 8883 (default for MQTT)
client.connect(mqtt_broker,mqtt_port)
# setting callbacks, use separate functions like above for better visibility
client.on_message = on_message

client.loop_start()

# Mở file Excel
wb = openpyxl.load_workbook('line_point.xlsx')
# Chọn sheet bạn muốn đọc
sheet = wb['sheetdata']  
while(1):
    for row in sheet.iter_rows(values_only=True,min_row=3,max_row=sheet.max_row):
        current_time = datetime.now().strftime('%Y-%m-%d  %H:%M:%S')
        default_value = '{"Real_Time":"%s","latitude":%f,"longitude":%f,"speed":%f,"direction":%f}'%(current_time,float(row[1]),float(row[2]),20,20672)
        client.publish(mqtt_topic,payload=default_value)
        time.sleep(2)

# '{"Real_Time":"2024-07-05  18:47:12","latitude":21.052468,"longitude":105.889678,"speed":20.00000=0,"direction":206.720000}' 
