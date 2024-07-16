import tkinter as tk
from tkinter import *
import tkintermapview 
from tkinter import ttk
from PIL import ImageTk, Image
from datetime import datetime
from tkinter import filedialog, messagebox
from openpyxl import Workbook
import time 
import threading
from threading import Lock
from MQTT_class import*
from Play_audio import*
import haversine as hs   
from haversine import Unit
import json 
import tkintermapview
# from data import distance_target
import openpyxl
import os


#---------------- initializing the variables    -----------------------
#Using queue to getdata from thread gps 
# Khóa dữ liệu để tránh xung đột
data_lock = Lock()
# Khởi tạo một đối tượng Lock
sound_lock = Lock()
# Cờ frame thông báo nhận tín hiệu
flag_frame1 = False
flag_frame2 = False
flag_frame3 = False

# Realtime data latitude - longitude from cloud
latitude_values=0.0
longitude_values=0.0
speed_values=0.0
direction = ""
#Location after transfer Latitude - Longitude to Addresss
Location=""
# Train status  
train_status="An Toàn"

# Location target
Latitude_target=0.0
Longitude_target=0.0

# target_id 
target_id=1
# Real time 
current_time=""
# values get to table
data=""
speed_id=""
# Train status  
train_status="An Toàn"
speed_warning=0 
# Turn ON / OFF notification 
turn_values=0
# Khởi tạo biến để lưu giá trị reference 
column_values = []
column_speed  = []
index_of_train = 4

#=========================================   connect to  MQTT   ===========================================

# /*MQTT Broker Connection Details*/
mqtt_broker = "ae501b5ee3194ca682bd67e257459478.s1.eu.hivemq.cloud"
mqtt_username = "RainWay System"
mqtt_password = "012301230123aA#"
mqtt_port = 8883
mqtt_topic_1 = "Train/Location_data"

#Using queue to getdata from thread gps 
data_queue = queue.Queue()

class MQTTClient():
    def __init__(self, broker, port, username, password, topic):
        self.broker = broker
        self.port = port
        self.username = username
        self.password = password
        self.topic = topic
        # client_id is the given name of the client 
        self.client = paho.Client(paho.CallbackAPIVersion.VERSION2)
        self.client.tls_set(tls_version=mqtt.client.ssl.PROTOCOL_TLS)
        self.client.on_connect = self.on_connect
        self.client.on_message = self.on_message
        self.client.username_pw_set(self.username, self.password)
        self.client.connect(self.broker, self.port)
        self.client.loop_start()
    def on_connect(self,client, userdata, flags, rc, properties=None):
        print("CONNACK received with code %s." % rc)
        self.client.subscribe(self.topic)
    # Callback function khi nhận được tin nhắn từ MQTT Broker
    def on_message(self,client, userdata, message):
        print("Received message '" 
            + str(message.payload.decode("utf-8")) 
            + "' on topic '"+ message.topic 
            + "' with QoS " + str(message.qos))
    def get_direction(true_course):
        if true_course is None:
            return "Unknown"
        directions = [
            "Bắc", "Bắc Đông Bắc", "Đông Bắc", "Đông Đông Bắc",
            "Đông", "Đông Đông Nam", "Đông Nam", "Nam Đông Nam",
            "Nam", "Nam Tây Nam", "Tây Nam", "Tây Tây Nam",
            "Tây", "Tây Tây Bắc", "Tây Bắc", "Bắc Tây Bắc"
        ]
        idx = int((true_course + 11.25) % 360 / 22.5)
        return directions[idx]
    # Publish data to hivemq
    def send_data_to_hivemq(self,time, speed, latitude, longitude, direction):
    # Tạo một dictionary chứa thông số cần gửi
        data = {
            "time": time,
            "speed": speed,
            "latitude": latitude,
            "longitude": longitude,
            "direction": direction
        }
        # Chuyển đổi dictionary thành chuỗi JSON
        json_data = json.dumps(data)
        # Đẩy dữ liệu lên topic MQTT
        self.client.publish(self.topic, json_data)
    def start(self):
        self.client.loop_start()
    def close_hivemq(self):
        self.client.disconnect()

# ================================================================================================================================
def parse_gps(line):
    try:
        # Kiểm tra xem dòng dữ liệu là thông điệp GGA, RMC hoặc GSV không
        if line.startswith('$GPGGA') or line.startswith('$GPRMC'):  
            msg = pynmea2.parse(line)
            #print(msg)
            if isinstance(msg, pynmea2.types.talker.RMC) and msg.latitude !=0.0 and msg.longitude !=0.0:
                data_queue.put(msg)
            while not data_queue.empty():
                # def update_speed(self):
                data = data_queue.get()
                global latitude_values,Longitude_values,speed_values,direction,Location,current_time
                direction = mqtt_client.get_direction(data.true_course) if data.true_course is not None else "Unknown"
                speed_values = data.spd_over_grnd
                adr = tkintermapview.convert_coordinates_to_address(data.latitude,data.longitude)
                Location = str(adr.street)+"\n"+str(adr.city) +"\n"+str(adr.country)
                Latitude_values = data.latitude
                Longitude_values= data.longitude
                #push data to hivemq
                mqtt_client.send_data_to_hivemq(str(current_time), str(Speed), str(Latitude_values), str(Longitude_values),data.true_course)
                
    except pynmea2.ParseError as e:
        print(f"Parse error: {e}")   
    except KeyboardInterrupt:
        print("Dừng đọc dữ liệu GPS.")

def read_gps():
    serial_port = serial.Serial('/dev/ttyS0', 9600, timeout=5)  # Mở cổng serial
    # Gửi lệnh cấu hình tần số cập nhật của GPS module thành 3 giây (0.2 Hz)
    ubx_command = b"\xB5\x62\x06\x08\x06\x00\xB8\x0B\x01\x00\x01\x00\xD9\x41"
    serial_port.write(ubx_command)
    while True:
        try:
            line = serial_port.readline().decode('unicode_escape')  # Đọc dòng dữ liệu từ cổng serial
            parse_gps(line)  # Gọi hàm parse dữ liệu GPS
        except Exception as e:
            print("Exception in read_gps:", e)
        
# ==============================================================================================================

# =================================================== Play sound ==============================================
def playAudio_Train_Station():
    global latitude_values, longitude_values, turn_values
    while True:
        if turn_values == 1:
            current_position = (latitude_values, longitude_values)
            # Kiểm tra khoảng cách đến các điểm quan trọng
            if hs.haversine(current_position, (21.02439, 105.84122), unit=Unit.KILOMETERS) <= 1:
                playsound('Audio/Audio_HaNoi.wav')
            if hs.haversine(current_position, (20.94663, 106.33057), unit=Unit.KILOMETERS) <= 1:
                playsound('Audio/Audio_HaiDuong.wav')
            elif hs.haversine(current_position, (20.85596, 106.68736), unit=Unit.KILOMETERS) <= 1:
                playsound('Audio/Audio_HaiPhong.wav')
            else:
                time.sleep(0.5)  # Ngủ khi không cần kiểm tra lại
        else:
            time.sleep(0.5)  # Ngủ khi không cần kiểm tra lại
# def playAudio_Train_Station():
#     global latitude_values, longitude_values, turn_values
#     # Danh sách điểm quan trọng
#     important_points = [
#         (21.02439, 105.84122, 'Audio/Audio_HaNoi.wav'),
#         (20.94663, 106.33057, 'Audio/Audio_HaiDuong.wav'),
#         (20.85596, 106.68736, 'Audio/Audio_HaiPhong.wav')
#     ]

#     # Tập hợp để theo dõi các điểm đã được thông báo
#     notified_points = set()
#     while True:
#         if turn_values == 1:
#             current_position = (latitude_values, longitude_values)
#             for point in important_points:
#                 if hs.haversine(current_position, (point[0], point[1]), unit=Unit.KILOMETERS) <= 1 and point not in notified_points:
#                     playsound(point[2])
#                     notified_points.add(point) 
#                     # Kiểm tra nếu tất cả các điểm đã được thông báo
#                     if len(notified_points) == len(important_points):
#                         print("Tất cả các điểm đã được thông báo. Dừng lại.")
#                         notified_points.clear()
#             time.sleep(0.5)  # Ngủ khi không cần kiểm tra lại
#         else:
#             time.sleep(0.5)  # Ngủ khi không cần kiểm tra lại
# def playAudio_Station():
#     global index_of_train
#     while True:
#         try:
#             if turn_values == 1:
#                 # Kiểm tra khoảng cách đến các điểm quan trọng
#                 if index_of_train < 2:
#                     playsound('Audio/Audio_HaNoi.wav')
#                 elif  233< index_of_train <= 235:
#                     playsound('Audio/Audio_HaiDuong.wav')
#                 elif 412 < index_of_train <= 414:
#                     playsound('Audio/Audio_HaiPhong.wav')
#                 else:
#                     time.sleep(1)  # Ngủ khi không cần kiểm tra lại
#             else:
#                 time.sleep(1)
#         except Exception as e:
#                 print(f"Play Audio in Station: {e}")
                         
def playAudio_Speed_Warning():
    while True:
        try:
            global column_values, speed_values, train_status, speed_warning
            speed_warning = column_speed[index_of_train]
            if speed_values > column_speed[index_of_train]:
                train_status = "Không an toàn"
                playsound('Audio/Audio_Canhbaotocdo.wav')
                time.sleep(1)  # Ngủ khi không cần kiểm tra lại
            else:
                train_status = "An toàn"
                time.sleep(1)  # Ngủ khi không cần kiểm tra lại
        except Exception as e:
                print(f"Play Audio Speed warning: {e}")
# ===============================================================================================================

# =====================================================  Main App  ==============================================

class RootApplication(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Train System")
        self.geometry("1280x720+150+20")
        self.resizable(False,False)
        self.configure(bg='gray')
        self.title("Hệ Thống Tàu Hỏa")
        # Icon 
        img = PhotoImage(file='image/eye.png')
        self.tk.call('wm', 'iconphoto', self._w, img)
        # BackGround:
        self.image = Image.open(r"image/Bground-min.png")
        self.img = ImageTk.PhotoImage(self.image)
        self.label_Bg=tk.Label(self,image=self.img)
        self.label_Bg.place(x=0,y=0) 
        # Khởi tạo và bắt đầu luồng phát âm thanh khoảng cách tàu và ga 
        self.speed_Warning_thread = threading.Thread(target=playAudio_Speed_Warning, daemon=True)
        self.speed_Warning_thread.start()
        # self.station_thread = threading.Thread(target=playAudio_Station, daemon=True)
        # self.station_thread.start()
        self.sound_thread = threading.Thread(target=playAudio_Train_Station, daemon=True)
        self.sound_thread.start()
        
        self.gps_thread =threading.Thread(target=read_gps)
        self.gps_thread.start()
        # Khởi tạo MQTT Client
        # self.mqtt_client = MQTTClient(mqtt_broker, mqtt_port, mqtt_username, mqtt_password, mqtt_topic_1)
        self.frames = {}  # Dictionary to store frames
        # Initialize frames
        self.create_frames()
        self.label_time = tk.Label(self, font=('Arial', 15, 'bold'), bg="white", fg="#4660ac", borderwidth=0, 
                                   border=0, width=18, relief='groove', justify=CENTER)
        self.label_time.place(x=54, y=102)
        #--------------------------- Func -------------------------------
        self.update_time()
         # Gán hàm xử lý sự kiện khi đóng cửa sổ
        self.protocol("WM_DELETE_WINDOW", self.on_close)

    def on_close(self):
        # Đóng kết nối MQTT
        mqtt_client.close_hivemq()

        # Chờ cho luồng phát âm thanh kết thúc
        self.speed_Warning_thread.join()
        # self.station_thread.join()
        self.sound_thread.join()
        self.gps_thread.join()
        pygame.mixer.quit()
        pygame.quit()

        # Hủy ứng dụng Tkinter
        self.destroy()

    def create_frames(self):
        #------------------------- frame2 -------------------------------
        self.frame2 = Frame2(self)
        self.frames["frame2"] = self.frame2
        self.frame2.config(bg="white", width=1180, height=550)
        self.frame2.place(x=52, y=146)
        self.button_frame2 = tk.Button(self, text="Theo dõi tàu", font=('Arial', 13, 'bold'), bg="white", fg="#4660ac", 
                                       activebackground="white", cursor='hand2', borderwidth=0, border=0, width=15, 
                                       relief='groove', justify=CENTER, command=lambda: self.raise_frame("frame2"))
        self.button_frame2.place(x=881, y=100)

        #------------------------- frame3 -------------------------------
        self.frame3 = Frame3(self)
        self.frames["frame3"] = self.frame3
        self.frame3.config(bg="white", width=1180, height=550)
        self.frame3.place(x=52, y=146)
        self.button_frame3 = tk.Button(self, text="Trang dữ liệu", font=('Arial', 13, 'bold'), bg="white", fg="#4660ac", 
                                       activebackground="white", cursor='hand2', borderwidth=0, border=0, width=15, 
                                       relief='groove', justify=CENTER, command=lambda: self.raise_frame("frame3"))
        self.button_frame3.place(x=1081, y=100)

        #------------------------- Frame1 -------------------------------
        self.frame1 = Frame1(self)
        self.frames["frame1"] = self.frame1
        self.frame1.config(bg="white", width=1180, height=550)
        self.frame1.place(x=52, y=146)
        self.button_frame1 = tk.Button(self, text="Trang thông tin", font=('Arial', 13, 'bold'), bg="white", fg="#4660ac", 
                                       activebackground="white", cursor='hand2', borderwidth=0, border=0, width=15, 
                                       relief='groove', justify=CENTER, command=lambda: self.raise_frame("frame1"))
        self.button_frame1.place(x=680, y=100)

    def raise_frame(self, frame_name):
        frame = self.frames.get(frame_name)
        if frame and frame.winfo_exists():
            frame.tkraise()
        else:
            print(f"{frame_name} does not exist or has been destroyed, recreating...")
            self.recreate_frame(frame_name)
            frame = self.frames[frame_name]
            frame.tkraise()

    def recreate_frame(self, frame_name):
        if frame_name == "frame1":
            self.frame1 = Frame1(self)
            self.frames["frame1"] = self.frame1
            self.frame1.config(bg="white", width=1180, height=550)
            self.frame1.place(x=52, y=146)
        elif frame_name == "frame2":
            self.frame2 = Frame2(self)
            self.frames["frame2"] = self.frame2
            self.frame2.config(bg="white", width=1180, height=550)
            self.frame2.place(x=52, y=146)
        elif frame_name == "frame3":
            self.frame3 = Frame3(self)
            self.frames["frame3"] = self.frame3
            self.frame3.config(bg="white", width=1180, height=550)
            self.frame3.place(x=52, y=146)
        #--------------------------- Func -------------------------------
        self.update_time()
    def update_time(self):
        # Real time """datetime(year, month, day[, hour[, minute[, second[, microsecond[,tzinfo]]]]])\
        global current_time
        current_time = datetime.now().strftime('%Y-%m-%d  %H:%M:%S')
        self.label_time.config(text=current_time)
        self.label_time.after(1000,self.update_time)
        
#==================================================== Frame 1: Trang điều khiển xe ===================================================
        
class Frame1(tk.Frame):
    def __init__(self, master):
        super().__init__(master)
        #-------------------------------- Train status -----------------------------------------------------------------------------
        self.label_status=Label(self,text="Trạng thái Tàu",fg='white',bg='#f22f2f',relief='groove',
                        width=20,borderwidth=1,border=1,justify=CENTER,font=("Arial", 15, "bold"),pady=5)
        self.label_status.place(x=29,y=160)
        self.label_Noti=Label(self,text="Trạng thái",fg='white',bg='#777777',relief='groove',pady=4,
                        width=17,borderwidth=1,border=1,justify=CENTER,font=("Arial", 13, "bold"))
        self.label_Noti.place(x=62,y=215)
        #--------------------------------------------------- Direction --------------------------------------------------------------
        self.label_direction=Label(self,text="Hướng di chuyển",fg='white',bg='#f22f2f',relief='groove',
                        width=20,borderwidth=1,border=1,justify=CENTER,font=("Arial", 15, "bold"),pady=5)
        self.label_direction.place(x=29,y=270)
        self.label_show=Label(self,text="hướng di chuyển",fg='white',bg='#777777',relief='groove',pady=4,
                        width=17,borderwidth=1,border=1,justify=CENTER,font=("Arial", 12, "bold"))
        self.label_show.place(x=62,y=325)
        #------------------------------------ show Location - Speed ----------------------------------------------------------------
        # Header label
        self.header1= Label(self,text="Địa chỉ hiện tại",fg='white',bg="#4660ac",width=30,borderwidth=0,border=1,justify=CENTER,
                       font=("Arial", 15, "bold"),relief='groove',pady=5)
        self.header1.place(x=400,y=110)
        self.header1.config(justify='center')
        # lable Location
        self.local1= Label(self,text="Địa chỉ",fg="#4660ac",width=30,borderwidth=1,border=1,justify=CENTER,
                           font=("Arial", 15, "bold"),pady=5)
        self.local1.place(x=400,y=165)
        #header speed
        self.header2= Label(self,text="Tốc độ hiện tại",fg='white',bg="#4660ac",width=30,borderwidth=0,border=1,justify=CENTER,
                       font=("Arial", 15, "bold"),relief='groove',pady=5)
        self.header2.place(x=400,y=270)
        # lable Speed
        self.speed1= Label(self,text="Tốc độ",fg="#4660ac",width=30,borderwidth=1,border=1,justify=CENTER,
                           font=("Arial", 15, "bold"),pady=5)
        self.speed1.place(x=400,y=325)
         #-------------------------------------- Target train --------------------------------------------------------------------------
        self.Label_target=Label(self,text="Mục tiêu di chuyển",fg='white',bg="#4660ac",width=30,borderwidth=0,border=1,justify=CENTER,
                       font=("Arial", 15, "bold"),pady=5)
        self.Label_target.place(x=400,y=380)
        self.Entry_Lat_Long_Target=Entry(self,fg="#4660ac",bg="#EDEBEB",relief='flat',width=19,borderwidth=1,border=1,justify=CENTER
                                         ,font=("Arial", 15, "bold"))
        self.Entry_Lat_Long_Target.place(x=550,y=435)        
        self.Label_target=Label(self,text="Vĩ độ/ Kinh độ",fg='white',bg="#4660ac",width=12,borderwidth=0,border=0,justify=CENTER,
                                        font=("Arial", 14, "bold"))
        self.Label_target.place(x=400,y=435)
        # --------------------------------------- Status Speed ---------------------------------------------------------------------------
    
        self.header_speed=Label(self,text="Cảnh báo tốc độ",fg='white',bg='#f22f2f',relief='groove',
                        width=20,borderwidth=1,border=1,justify=CENTER,font=("Arial", 15, "bold"),pady=5)
        self.header_speed.place(x=900,y=160)
        self.header_speed_noti = Label(self,text="Cảnh báo",fg='white',bg='#777777',relief='groove',pady=4,
                        width=17,borderwidth=1,border=1,justify=CENTER,font=("Arial", 12, "bold"))
        self.header_speed_noti.place(x=933,y=215)
        # -------------------------------------------------- TURN ON /OFF notification ----------------------------------------------------
        # Change Turn values:
        self.turn_values_tk = tk.IntVar()
        self.turn_values_tk.set(1)
        self.check_turn =tk.Checkbutton(self,text = "Âm thanh thông báo", variable = self.turn_values_tk,font=("Arial", 15, "bold"),
                                        fg="black",bg="#f22f2f",cursor='hand2',width=18,justify=CENTER) 
        self.check_turn.place(x=900,y=270)
        self.update_frame1_realtime()
        self.update_label_location_target()

    def update_label_location_target(self):
        global Latitude_target,Longitude_target,target_id,turn_values
        turn_values =self.turn_values_tk.get()
        if target_id==0:
            self.Entry_Lat_Long_Target.delete(0,'end')
            Latitude_target= None
            Longitude_target= None
            target_id=1
        data_target = self.Entry_Lat_Long_Target.get() 
        if data_target !="" and data_target is not None:
            parts = data_target.split()
            try:  # Kiểm tra xem danh sách có phần tử nào không trước khi truy cập
                Latitude_target=float(parts[0])
                Longitude_target=float(parts[1])
            except Exception as e:
                print(f"An error occurred: {e}")
        self.after(1000,self.update_label_location_target)
        
    def update_frame1_realtime(self):
        global train_status,Location,speed_values,longitude_values,latitude_values,direction,flag_frame1,speed_warning
        # self.update_label_location_target()
        if flag_frame1:
            flag_frame1 = False
            try:    
                self.header_speed_noti.config(text="%d Km/h"%speed_warning)
                self.label_Noti.config(text=train_status)
                self.local1.config(text=Location)
                self.label_show.config(text=direction)
                self.speed1.config(text="%d Km/h"%speed_values)
            except Exception as e:
                print(f"An error occurred in frame1 : {e}")
        self.after(500, self.update_frame1_realtime)

#------------------------- Frame 2:  Show map -----------------------------------------------------------
        
class Frame2(tk.Frame):
    def __init__(self, master):
        super().__init__(master)
        
        #------------------- header location detail ---------------------------------------------------
        # Header latitude
        self.header2_1= Label(self,text="vĩ Độ",fg='white',bg="#4660ac",width=12,borderwidth=0,border=1,justify=CENTER
                              ,font=("Arial", 12, "bold"),relief='groove',pady=2)
        self.header2_1.place(x=0,y=180)
        # Header Longitude
        self.header2_2= Label(self,text="Kinh Độ",fg='white',bg="#4660ac",width=12,borderwidth=0,border=1,justify=CENTER
                              ,font=("Arial", 12, "bold"),relief='groove',pady=2)
        self.header2_2.place(x=0,y=250)
        # Header speed
        self.header2_3= Label(self,text="Tốc độ",fg='white',bg="#4660ac",width=12,borderwidth=0,border=1,justify=CENTER
                              ,font=("Arial", 12, "bold"),relief='groove',pady=2)
        self.header2_3.place(x=0,y=320)
        # Header address
        self.header2_4= Label(self,text="Địa chỉ",fg='white',bg="#4660ac",width=15,borderwidth=0,border=1,justify=CENTER
                              ,font=("Arial", 12, "bold"),relief='groove',pady=2)
        self.header2_4.place(x=1032,y=180)
        # Header distance: 
        self.header2_5= Label(self,text="Khoảng cách",fg='white',bg="#4660ac",width=15,borderwidth=0,border=1,justify=CENTER
                              ,font=("Arial", 12, "bold"),relief='groove',pady=2)
        self.header2_5.place(x=1032,y=280)
        # Header direction 
        self.header2_6  =Label(self,text="Hướng di chuyển",fg='white',bg="#4660ac",width=15,borderwidth=0,border=1,justify=CENTER
                               ,font=("Arial", 12, "bold"),relief='groove',pady=2)
        self.header2_6.place(x=1032,y=350)
        #------------------- values location details ---------------------------------------------------
        # latitude values
        self.latitude= Label(self,text="Vĩ Độ",fg='Gray',width=15,borderwidth=0,border=1,justify=CENTER
                             ,font=("Arial", 10, "bold"),relief='groove',pady=2)
        self.latitude.place(x=0,y=210)
        # Longitude values
        self.Longitude= Label(self,text="Kinh Độ",fg='Gray',width=15,borderwidth=0,border=1,justify=CENTER
                              ,font=("Arial", 10, "bold"),relief='groove',pady=2)
        self.Longitude.place(x=0,y=280)
        # speed values
        self.speed= Label(self,text="Tốc Độ",fg='Gray',width=15,borderwidth=0,border=1,justify=CENTER
                          ,font=("Arial", 10, "bold"),relief='groove',pady=2)
        self.speed.place(x=0,y=350)
         # address values
        self.Address= Label(self,text="Địa chỉ",fg='Gray',width=21,borderwidth=0,border=1,justify=CENTER
                            ,font=("Arial", 9, "bold"),relief='groove',pady=2)
        self.Address.place(x=1032,y=210)
         # Distance values
        self.Distance= Label(self,text="Khoảng cách",fg='Gray',width=21,borderwidth=0,border=1,justify=CENTER
                             ,font=("Arial", 9, "bold"),relief='groove',pady=2)
        self.Distance.place(x=1032,y=310)
        # direction values
        self.Direction= Label(self,text="Hướng đi",fg='Gray',width=21,borderwidth=0,border=1,justify=CENTER
                              ,font=("Arial", 9, "bold"),relief='groove',pady=2)
        self.Direction.place(x=1032,y=380)
        #---------------------Map view-------------------------------------------------------

        self.map_widget = tkintermapview.TkinterMapView(self, width=900, height=550, corner_radius=0)
        self.map_widget.place(relx=0.49, rely=0.5, anchor=CENTER)
        # self.map_widget.set_tile_server("https://a.tile.openstreetmap.org/{z}/{x}/{y}.png")  # OpenStreetMap (default)
        self.map_widget.set_tile_server("https://mt0.google.com/vt/lyrs=m&hl=en&x={x}&y={y}&z={z}&s=Ga", max_zoom=22)  # google normal
        # self.map_widget.set_tile_server("https://mt0.google.com/vt/lyrs=s&hl=en&x={x}&y={y}&z={z}&s=Ga", max_zoom=22)  # google satellite
        self.marker1 =None
        # self.marker2 =None
        global column_values
        # for row in sheet.iter_rows(values_only=True, min_row=3, max_row=sheet.max_row-1):
        #     # print(row[1],row[2])
        #     column_values.append((float(row[1]), float(row[2])))
        path_1= self.map_widget.set_path(column_values)
            
        #----------------------Show location - Real time------------------------
        self.map_widget.add_right_click_menu_command(label="xóa mục tiêu",command=self.clear_marker_event,
                                                pass_coords=True)
        self.update_location_map()
        self.update_location_target()

    # update location in map
    def update_location_map(self):
        global latitude_values,longitude_values,train_status,speed_values,direction,Location,flag_frame2
        if flag_frame2:
            flag_frame2 = False
            # print(latitude_values,longitude_values)
            self.Address.config(text=Location)
            # update_label_direction 
            self.Direction.config(text=direction)
            # update_speed()
            self.speed.config(text="%d Km/h"%speed_values)
            self.latitude.config(text=latitude_values)
            self.Longitude.config(text=longitude_values)
            try:
                if self.marker1 is not None:
                    self.marker1.delete()
                self.marker1= self.map_widget.set_position(latitude_values, longitude_values,text="Vị trí tàu",marker=True)
            except Exception as e:
                print(e)
        else:
            pass
        # Update the position on the map
        self.after(500,self.update_location_map)
    def update_location_target(self):
        global latitude_values,longitude_values,target_id,Latitude_target,Longitude_target,turn_values
        if Latitude_target is not None and Longitude_target is not None:
            try:
                current_position=(latitude_values,longitude_values)
                target_position =(Latitude_target,Longitude_target)
                distance_2_point=self.distance_target(current_position,target_position)
                self.marker2=self.map_widget.set_marker(Latitude_target,Longitude_target,text="Điểm mục tiêu")
                # update distance values
                # self.Distance.config(text=str(distance_2_point))
                self.Distance.config(text="%f Km"%distance_2_point)
                # Load file âm thanh
                if distance_2_point <= 1.0 :
                    target_id=0
                    self.map_widget.delete_all_marker()
                    if turn_values == 1:
                        playsound('Audio/Audio_MucTieu.wav')
                    self.Distance.config(text="Hoàn thành")
                # mutex.release()
            except Exception as e:
                print(f"Location Target in Map:{e}")
                target_id=1
        self.after(1000,self.update_location_target)
    def clear_marker_event(self,event=None):
        global target_id 
        try:
            target_id=0
            self.map_widget.delete_all_marker()
            self.Distance.config(text="")
        except Exception as e:
            print(f"An error occurred in frame2: {e}")
    # Mở file Excel
    def distance_target(self,current_position,target_position): 
        global column_values     
        # Tìm điểm trên dải tọa độ gần nhất
        nearest_coord = min(column_values, key=lambda coord: hs.haversine(current_position, coord))
        # Xác định chỉ mục của điểm gần nhất trong danh sách line_coords
        index_of_train = column_values.index(nearest_coord)

        # Tìm điểm trên dải tọa độ gần nhất
        target_coord = min(column_values, key=lambda coord: hs.haversine(target_position, coord))
        # Xác định chỉ mục của điểm gần nhất trong danh sách line_coords
        index_of_target = column_values.index(target_coord)
        total =0.0
        for i in range(min(index_of_train, index_of_target),max(index_of_train, index_of_target)):
            distance = hs.haversine(column_values[i],column_values[i + 1])
            total+=distance
        return total 
           
#----------------------------------------------------- Frame 3: Data Receive ---------------------------------------------------------------------
        
class Frame3(tk.Frame):
    def __init__(self, master):
        super().__init__(master)
        # Tạo Treeview
        self.tree = ttk.Treeview(self, columns=("Time", "Latitude", "Longtitud","Speed","Location","Status"), show="headings")
        self.tree.place(width=1000,height=500,x=0,y=20)
        self.tree.heading("Speed", text="Tốc độ")
        self.tree.heading("Latitude", text="Vĩ độ")
        self.tree.heading("Longtitud", text="Kinh độ")
        self.tree.heading("Location", text="Địa chỉ hiện tại")
        self.tree.heading("Status", text="Trạng thái")
        self.tree.heading("Time", text="Thời gian")
        # Thiết lập chiều rộng cột
        self.tree.column("Speed",anchor="center", width=30)
        self.tree.column("Latitude",anchor="center", width=50)
        self.tree.column("Longtitud", anchor="center",width=50)
        self.tree.column("Location", anchor="center",width=250)
        self.tree.column("Status", anchor="center",width=50)
        self.tree.column("Time", anchor="center",width=65)
        # Tạo scrollbar cho treeview
        self.scrollbar1 = ttk.Scrollbar(self, orient="vertical", command=self.tree.yview)
        self.scrollbar2 = ttk.Scrollbar(self, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=self.scrollbar1.set,xscrollcommand=self.scrollbar2.set)
        self.scrollbar1.place(height=500,x=1000,y=20)
        self.scrollbar2.place(width=1000,x=0,y=520)
        #------------------------- get data to xlsx file ------------------------------------------
        self.button_save = tk.Button(self,text="xuất dữ liệu",padx=8,pady=3,background="#4660ac",fg="white",font=("Arial", 15, "bold")
                                     ,border=0,borderwidth=0,justify=CENTER,cursor='hand2',command=self.get_data)
        self.button_save.place(x=1030,y=250)
        self.update_data()
        self.schedule_backup()  # Bắt đầu sao lưu định kỳ
    # update data to table:
    def update_data(self):
        # Dữ liệu mẫu
        global current_time,Location,latitude_values,longitude_values,train_status,speed_values,flag_frame3
        # Loại bỏ ký tự xuống dòng (\n)
        if flag_frame3:
            flag_frame3 = False
            Location_no_newline = Location.replace("\n", "")
            data = [str(current_time), str(latitude_values), str(longitude_values),str(speed_values)+" km/h",Location_no_newline,train_status]
            self.tree.insert("", "end", values=data)
        self.after(250,self.update_data)      
    def get_data(self):
        # Tạo một workbook mới
        self.wb = Workbook()
        # Chọn active worksheet (sheet mặc định)
        self.ws = self.wb.active
        # Ghi dữ liệu vào worksheet
        self.ws.append(["Time", "Latitude", "Longtitude","Speed","Location","Status"])
        for item in self.tree.get_children():
            self.values = self.tree.item(item, 'values')
            self.ws.append(self.values)
        
        # Hiển thị hộp thoại để chọn vị trí lưu tệp
        file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        if file_path:
            # Lưu workbook thành file Excel tại đường dẫn đã chọn
            try:
                self.wb.save(file_path)
                self.clear_data_received()
                self.delete_backup()  # Xóa backup khi dữ liệu được export thành công
            except Exception as e:
                messagebox.showerror("Lỗi", f"Có lỗi xảy ra khi lưu file: {e}")
    def clear_data_received(self):
        for item in self.tree.get_children(): # used self.tree instead
            self.tree.delete(item)
    def backup_data(self):
        backup_file_path = 'backup_data.xlsx'
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.append(["Time", "Latitude", "Longitude", "Speed", "Location", "Status"])
        for item in self.tree.get_children():
            self.values = self.tree.item(item, 'values')
            self.ws.append(self.values)
        try:
            self.wb.save(backup_file_path)
            print("Backup saved successfully.")
        except Exception as e:
            print(f"Error while saving backup: {e}")

    def delete_backup(self):
        backup_file_path = 'backup_data.xlsx'
        if os.path.exists(backup_file_path):
            os.remove(backup_file_path)
            print("Backup file deleted.")
    def schedule_backup(self):
        self.backup_data()
        self.master.after(300000, self.schedule_backup)  # Sao lưu mỗi 5 phút
            
if __name__ == "__main__":
    #Các điểm ga tàu - lưu trữ 
    wb = openpyxl.load_workbook('line_point.xlsx')
    sheet = wb['sheetdata']  # Thay 'sheetdata' bằng tên sheet cần đọc
    # Đọc các giá trị từ sheet và tính khoảng cách
    for row in sheet.iter_rows(values_only=True, min_row=3, max_row=sheet.max_row-1):
        column_values.append((float(row[1]), float(row[2])))
        column_speed.append((float(row[3])))
    mqtt_client = MQTTClient(mqtt_broker, mqtt_port, mqtt_username, mqtt_password, mqtt_topic_1)
    app = RootApplication()
    app.mainloop()